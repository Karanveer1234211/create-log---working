#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
journal_to_excel.py  -  Human-readable historical Excel from the prediction journal.

Turns prediction_journal.parquet into a wide, eyeball-friendly workbook like a
hand-kept log, but richer: entry + the 5 forward closes + movement, the model's
1/3/5-day outcomes, MFE/MAE, event flags, and OBJECTIVE structural proxies at
short / medium / long horizons (distance from 20d / 100d / 52w highs, and position
vs the 50- and 200-day moving averages). These proxies are the computable stand-in
for "near a trendline" - real numbers, point-in-time, no hand-drawn lines.

Two sheets: OOS (genuine out-of-sample, your real evidence) and InSample
(model trained on these - contaminated; included only for contrast, flagged).

IMPORTANT: this is a VIEWING tool. Eyeballing risers vs fallers against these
columns generates hunches, not conclusions - the denominator-respecting answer is
'prediction_journal.py --mode analyze'. Do not hand-tune training off this sheet.

USAGE:
  python journal_to_excel.py --journal "...\prediction_journal.parquet" ^
      --panel "...\out_rank_single\panel_cache.parquet" ^
      --out "...\prediction_journal.xlsx"   [--oos-only]

Self-test:  python journal_to_excel.py --self-test
"""
import argparse
import sys
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PCT = "0.0%"
PRICE = "#,##0.00"
COLUMNS = [
    ("pred_date", "Pred date", "yyyy-mm-dd"),
    ("symbol", "Symbol", None),
    ("rank", "Rank", "0"),
    ("prob", "Model prob", PCT),
    ("entry_ref", "Model last price", PRICE),
    ("entry_open", "Entry (next open)", PRICE),
    ("close_d1", "Close d1", PRICE),
    ("close_d2", "Close d2", PRICE),
    ("close_d3", "Close d3", PRICE),
    ("close_d4", "Close d4", PRICE),
    ("close_d5", "Close d5", PRICE),
    ("movement", "Movement", PRICE),
    ("pct_movement", "% Movement (5d)", PCT),
    ("ret_1", "Ret 1d", PCT),
    ("ret_3", "Ret 3d", PCT),
    ("mfe_5", "MFE 5d", PCT),
    ("mae_5", "MAE 5d", PCT),
    ("ev_gap", "Gap?", "0"),
    ("ev_vol_spike", "VolSpike?", "0"),
    ("ev_large_move", "BigMove?", "0"),
    ("dist_20d_high", "Dist 20d-high", PCT),
    ("dist_100d_high", "Dist 100d-high (~wk)", PCT),
    ("dist_252d_high", "Dist 52w-high (~mo)", PCT),
    ("pos_50dma", "vs 50DMA", PCT),
    ("pos_200dma", "vs 200DMA", PCT),
    ("stock_regime", "Regime", None),
]


def _pick(cols, *cands):
    low = {c.lower(): c for c in cols}
    for c in cands:
        if c in low:
            return low[c]
    return None


def load_panel(path):
    df = pd.read_parquet(path)
    d = _pick(df.columns, "timestamp", "date")
    s = _pick(df.columns, "symbol", "ticker")
    df = df.rename(columns={d: "date", s: "symbol"})
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.tz_localize(None).dt.normalize()
    return df.sort_values(["symbol", "date"]).reset_index(drop=True)


def panel_extras(panel):
    """Per-symbol: next-open entry, 5 forward closes, structural proxies (point-in-time)."""
    p = panel.copy()
    g = p.groupby("symbol", group_keys=False)
    p["entry_open"] = g["open"].shift(-1)
    p.loc[~(p["entry_open"] > 0), "entry_open"] = np.nan
    for k in range(1, 6):
        p[f"close_d{k}"] = g["close"].shift(-k)

    def roll(w, fn):
        return g["close"].transform(lambda s: getattr(s.rolling(w, min_periods=max(5, w // 5)), fn)())

    p["dist_20d_high"] = p["close"] / roll(20, "max") - 1
    p["dist_100d_high"] = p["close"] / roll(100, "max") - 1
    p["dist_252d_high"] = p["close"] / roll(252, "max") - 1
    p["pos_50dma"] = p["close"] / roll(50, "mean") - 1
    p["pos_200dma"] = p["close"] / roll(200, "mean") - 1
    keep = ["symbol", "date", "entry_open"] + [f"close_d{k}" for k in range(1, 6)] \
        + ["dist_20d_high", "dist_100d_high", "dist_252d_high", "pos_50dma", "pos_200dma"]
    return p[keep]


def build_table(journal, panel):
    j = journal.copy()
    j["pred_date"] = pd.to_datetime(j["pred_date"], errors="coerce").dt.tz_localize(None).dt.normalize()
    ex = panel_extras(panel).rename(columns={"date": "pred_date"})
    j = j.drop(columns=[c for c in ("entry_open",) if c in j.columns])
    m = j.merge(ex, on=["symbol", "pred_date"], how="left")
    m["movement"] = m["close_d5"] - m["entry_open"]
    m["pct_movement"] = m["close_d5"] / m["entry_open"] - 1
    for c in ("ret_1", "ret_3", "mfe_5", "mae_5"):     # journal stores in PERCENT -> fraction
        if c in m.columns:
            m[c] = pd.to_numeric(m[c], errors="coerce").replace([np.inf, -np.inf], np.nan) / 100.0
    if "prob" in m.columns:
        m["prob"] = pd.to_numeric(m["prob"], errors="coerce")
    for c, _, _ in COLUMNS:
        if c not in m.columns:
            m[c] = np.nan
    m["_is_oos"] = pd.to_numeric(m.get("is_oos", 1), errors="coerce").fillna(1).astype(int)
    return m


def _fmt_sheet(ws, header_row=1):
    headers = [h for _, h, _ in COLUMNS]
    fill = PatternFill("solid", start_color="1F3864")
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for j, (_, header, fmt) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(20, len(header) + 2))
        if fmt:
            for i in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=i, column=j).number_format = fmt


def write_sheet(ws, df, header_row=1):
    if header_row > 1:
        for _ in range(header_row - 1):
            ws.append([])
    ws.append([h for _, h, _ in COLUMNS])
    keys = [k for k, _, _ in COLUMNS]
    for _, r in df.iterrows():
        ws.append([r.get(k) for k in keys])
    _fmt_sheet(ws, header_row)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{ws.max_row}"


def export(journal_path, panel_path, out_path, oos_only=False):
    from openpyxl import Workbook
    journal = pd.read_parquet(journal_path)
    journal = journal[journal.get("status", "RESOLVED") == "RESOLVED"].copy()
    panel = load_panel(panel_path)
    tbl = build_table(journal, panel).sort_values(["pred_date", "rank"], ascending=[False, True])

    wb = Workbook()
    wb.remove(wb.active)
    oos = tbl[tbl["_is_oos"] == 1]
    write_sheet(wb.create_sheet("OOS"), oos)
    print(f"[xlsx] OOS sheet: {len(oos)} rows")
    if not oos_only:
        isr = tbl[tbl["_is_oos"] == 0]
        if len(isr):
            ws = wb.create_sheet("InSample (contaminated)")
            ws.cell(row=1, column=1, value="IN-SAMPLE - model trained on these dates. "
                    "Contaminated; for contrast only.").font = Font(bold=True, color="C00000")
            write_sheet(ws, isr, header_row=3)
            print(f"[xlsx] InSample sheet: {len(isr)} rows")
    wb.save(out_path)
    print(f"[xlsx] saved -> {out_path}")


def run_self_test():
    import tempfile
    rng = np.random.default_rng(3)
    syms = [f"S{i}" for i in range(25)]
    days = pd.bdate_range("2025-01-01", periods=400)
    rows = []
    for s in syms:
        px = rng.uniform(50, 800) * np.cumprod(1 + rng.normal(0.0005, 0.02, len(days)))
        for d, c in zip(days, px):
            rows.append(dict(symbol=s, timestamp=d, open=c * (1 + rng.normal(0, 0.003)),
                             high=c * 1.01, low=c * 0.99, close=c, volume=rng.uniform(1e5, 1e6)))
    panel = pd.DataFrame(rows)
    tmp = Path(tempfile.mkdtemp())
    panel.to_parquet(tmp / "panel.parquet")
    jr = []
    for di in range(50, 380, 7):
        for rank, s in enumerate(rng.choice(syms, 20, replace=False), 1):
            jr.append(dict(pred_date=days[di].date(), symbol=s, prob=round(rng.uniform(0.55, 0.9), 3),
                           rank=rank, entry_ref=100.0, ret_1=rng.normal(0, 2), ret_3=rng.normal(0, 3),
                           ret_5=rng.normal(0.5, 4), mfe_5=abs(rng.normal(3, 1)), mae_5=-abs(rng.normal(3, 1)),
                           ev_gap=0, ev_vol_spike=0, ev_large_move=int(rng.random() < 0.1),
                           is_oos=int(di > 200), stock_regime="bull_trend", status="RESOLVED"))
    pd.DataFrame(jr).to_parquet(tmp / "journal.parquet")
    out = tmp / "out.xlsx"
    export(str(tmp / "journal.parquet"), str(tmp / "panel.parquet"), str(out), oos_only=False)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "OOS" in wb.sheetnames
    ws = wb["OOS"]
    assert ws.max_column == len(COLUMNS) and ws.max_row > 10
    for row in ws.iter_rows(values_only=True):
        for v in row:
            assert not (isinstance(v, str) and v.startswith("#")), f"excel error {v}"
    print("\nSELF-TEST OK")
    return 0


def main():
    ap = argparse.ArgumentParser(description="Export prediction journal to historical Excel.")
    ap.add_argument("--journal"); ap.add_argument("--panel"); ap.add_argument("--out")
    ap.add_argument("--oos-only", dest="oos_only", action="store_true")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        sys.exit(run_self_test())
    if not (a.journal and a.panel and a.out):
        raise SystemExit("need --journal --panel --out")
    export(a.journal, a.panel, a.out, a.oos_only)


if __name__ == "__main__":
    main()
